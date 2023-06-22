# -*- coding: utf-8 -*-
"""
Created on Tue Aug 11 12:16:03 2020

@author: lvazquez

It determines, database based, the equipment models and client information to
extract the PQZ files from ELSPEC measurements equipments. With the client
information a folder tree is created to organize the files in a server,
creation date based.

Dteermina, con base en una base de datos, los modelos de los equipos y la
informacion de los clientes para extraer los archivos PQZ de equipos de
medicion ELSPEC. Con la informacion del cliente, un arbol de carpetas es
creado para organizar los archivos en un servidor, con base en su fecha de
creacion.

"""
# Modulos
import pyodbc
import sys
import datetime
import getpass
import openpyxl
import os

# Scripts externos de extraccion
from PureBBs import PureBBPQZs
from G44X0s import G44X0PQZs


# Funciones
def DBConnect(server='', database='', username='', password='', logFile=''):
    '''

    Parameters
    ----------
    server : string, Required
        DESCRIPTION. The default is ''. It is the IP direction of the remote
        database server where the client and equipment information is stored.
    database : string, Required
        DESCRIPTION. The default is ''. Name of the DB.
    username : string, Required
        DESCRIPTION. The default is ''. Username to acces the DB information.
    password : string, Required
        DESCRIPTION. The default is ''. Password related to the used username
        to acces the DB information.
    logFile : TextIOWrapper, Required
        DESCRIPTION. The default is ''. Log File Reference to write on.

    Returns
    -------
    conexion : Connection
        DESCRIPTION: DB connection reference.
    cursor : Cursor
        DESCRIPTION: DB cursor to point.

    '''
    if server == '' or password == '' or username == '' or database == '':
        NOW = datetime.datetime.now()  # Tiempo exacto actual
        NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                   '[%H:'+'%M:'+'%S]'))
        log = (NOW_log+'~'+'ERROR de conexion con BD:' +
               ' Un parametro no fue ingresado.')
        logFile.write(log + '\n')
        print(log)
        sys.exit(1)  # ERROR devuelve 0
    elif (type(server) != str or type(database) != str or
          type(username) != str or type(password) != str):
        NOW = datetime.datetime.now()  # Tiempo exacto actual
        NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                   '[%H:'+'%M:'+'%S]'))
        log = (NOW_log+'~'+'ERROR de conexion con BD:' +
               ' Los tipos de datos deben estar en formato string (texto)')
        logFile.write(log+'\n')
        print(log)
        sys.exit(1)  # Error devuelve 0
    else:
        print('Iniciando conexion a la base de datos')
        conexion = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};' +
                                  'SERVER='+server+';' +
                                  'DATABASE='+database+';' +
                                  'UID='+username+';' +
                                  'PWD=' + password)
        cursor = conexion.cursor()
        NOW = datetime.datetime.now()  # Tiempo exacto actual
        NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                   '[%H:'+'%M:'+'%S]'))
        log = NOW_log+'~'+'Conexion exitosa a la base de datos.'
        logFile.write(log+'\n')
        print(log)
        return conexion, cursor


def getDBdata(parametro='', tabla='', param_cond='', valor_cond='',
              logFile='', Query=''):
    '''

    Parameters
    ----------
    parametro : string, Required
        DESCRIPTION. The default is ''. It's the aremeter the user wants to get
        from the database table.
    tabla : string, Required
        DESCRIPTION. The default is ''. Name of the table where the user wants
        to get the values from.
    param_cond : string, Required
        DESCRIPTION. The default is ''. Known conditional parameter used as
        reference to get the desired paremeter.
    valor_cond : string, Required
        DESCRIPTION. The default is ''. Known conditional value of the
        conditional parameter.
    logFile : TextIOWrapper, Required
        DESCRIPTION. The default is ''. Log File Reference.
    Query : string, Optional
        DESCRIPTION. The default is ''. Parameter used when an specific Query
        is needed. Makes the function to ignore all other parameters and follow
        just this query.

    Returns
    -------
    rows : Row
        DESCRIPTION. Values obtained from the DB Query in row Type Tuples.

    '''
    if Query == '':
        if parametro == '' or tabla == '':
            NOW = datetime.datetime.now()  # Tiempo exacto actual
            NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                       '[%H:'+'%M:'+'%S]'))
            log = (NOW_log+'~'+'ERROR de consulta de BD: Un parametro de los' +
                   ' minimos obligatorios no fue llenado')
            logFile.write(log+'\n')
            print(log)
            # sys.exit(1)

        if param_cond != '' and valor_cond == '':  # Parametro: SI Valo: No
            NOW = datetime.datetime.now()  # Tiempo exacto actual
            NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                       '[%H:'+'%M:'+'%S]'))
            log = (NOW_log+'~'+"ERROR de consulta de BD: Si utiliza un" +
                   " parametro de condicion para 'WHERE'," +
                   " el valor de condicion es requerido")
            logFile.write(log+'\n')
            print(log)
            # sys.exit(1)

        if (type(parametro) != str or type(tabla) != str or
                type(param_cond) != str):
            NOW = datetime.datetime.now()  # Tiempo exacto actual
            NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                                       '[%H:'+'%M:'+'%S]'))
            log = (NOW_log+'~'+'ERROR de consulta de BD:' +
                   ' Los tipos de datos deben estar en formato string (texto)')
            logFile.write(log+'\n')
            print(log)
            # sys.exit()
        # print('Iniciando consulta...')
        # Generacion del QUERY si no hay condicion
        if param_cond == '':  # Todos los valores de una columna en una tabla
            Query = 'SELECT ' + parametro + ' FROM ' + tabla
        else:  # Cuando hay una condicion solo para ciertos parametros
            Query = ('SELECT '+parametro+' FROM '+tabla +
                     ' WHERE '+param_cond+'='+str(valor_cond))
        # print(Query)
        cursor.execute(Query)
        rows = cursor.fetchall()
    elif Query != '':
        print(Query)
        cursor.execute(Query)
        rows = cursor.fetchall()
    # print('Consulta exitosa!')
    return rows

    if parametro == '' or tabla == '':
        print('ERROR: Un parametro de los minimos obligatorios no fue llenado')
        sys.exit()

    if param_cond != '' and valor_cond == '':  # Si hay parametro pero no valor
        print("ERROR: Si utiliza un parametro de condicion para 'WHERE'," +
              " el valor de condicion es requerido")
        sys.exit()

    if type(parametro) != str or type(tabla) != str or type(param_cond) != str:
        # except:
        print('ERROR: Los tipos de datos deben estar' +
              ' en formato string (texto)')
        sys.exit()
    # print('Iniciando consulta...')
    # Generacion del QUERY si no hay condicion
    if param_cond == '':  # Todos los valores de una columna en una tabla
        Query = 'SELECT ' + parametro + ' FROM '+tabla
    else:  # Cuando hay una condicion solo para ciertos parametros
        Query = ('SELECT '+parametro+' FROM '+tabla +
                 ' WHERE '+param_cond+'='+str(valor_cond))
    # print(Query)
    cursor.execute(Query)
    rows = cursor.fetchall()
    # print('Consulta exitosa!')
    return rows


def logSave(logFile, m, debug='1'):
    '''

    Parameters
    ----------
    logFile : TextIOWrapper, Required
        DESCRIPTION. Text Document Reference..
    m : string, Required
        DESCRIPTION. Message that is going to be written into the Log File.
    debug : str, optional
        DESCRIPTION. The default is '1'. Varible used to determine if the
        messages are going to be writeen into the text file.

    Returns
    -------
    None.

    '''
    NOW = datetime.datetime.now()  # Tiempo exacto actual
    # Tiempor en formato deseado
    NOW_log = str(NOW.strftime('[%Y'+'-'+'%m'+'-'+'%d]' +
                               '[%H:'+'%M:'+'%S]'))
    log = NOW_log + '~' + m
    if debug == '1':
        logFile.write(log+'\n')
    print(log)


# Diccionario para ruta de log
logFileDict = {'logFileServerDev': 'S:\\Documents\\Python Scripts\\' +
               'TransferenciaAzure\\PQZLOG.txt',
               'logFileLocalDev': 'C:\\Users\\mjimenez\\' +
               'OneDrive - DIRAM SA DE CV\\Servicios\\Transferencia\\PQZLOG',
               'logFileProd': 'S:\\TransferenciaAProd\\PQZLOG.txt'}

# logFile = open(logFileDict['logFilePrueba'], 'a')
# logFile = open(logFileDict['logFileProd'], 'a')
logFile = open(logFileDict['logFileLocalDev'], 'a')

# °°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°
# ||||Obtencion de informacion de ips y modelos por medio de base de datos||||
try:
    # Parametros de conexion a la base de datos
    DBparams = {'serverDev': '172.16.104.36',
                'serverProd': '172.16.103.36',
                'DBname': 'BD_CLOUD_PQ',
                'userDev': 'luisvazquez',
                'userProd': 'mjimenez', #'PQZ',
                'passDev': 'vBJPKpEXnPpQy35E',
                'passProd': 'OcA2QlYv@2552p8!'}#'LadrilloVerde3!'}

    server = DBparams['serverProd']  # BD de produccion
    database = DBparams['DBname']
    username = DBparams['userProd']
    password = DBparams['passProd']
    # Conexion con la base de datos
    conexion, cursor = DBConnect(server,
                                 database,
                                 username,
                                 password,
                                 logFile)

    # Guardar parametro debug
    parametro = 'valor'
    tabla = 'Z01_Parametro'
    param_cond = 'descripcion'
    valor_cond = "'" + "Debug" + "'"

    try:  # Obtener el parametro debug en la base de datos
        debug = getDBdata(parametro,
                          tabla,
                          param_cond,
                          valor_cond,
                          logFile)
        debug = str(debug)
        debug = debug[11:12]  # Tomar solo el valor numerico en forma de string
    except Exception:
        debugv = '1'
        logSave(logFile,
                'Error en la consulta del parametro Debug para generacion' +
                ' de Log. Se usara valor predeterminado (On)')

    logSave(logFile, 'Conexion realizada de manera exitosa', debug)

    # Parametros para consulta de datos de los medidores activos
    # Consultamos ips y modelos de los medidores activos
    parametro = 'id_medidor_activo,ip_fija,folder'
    tabla = 'A08_Medidor_Activo'
    Query = ('SELECT top 1 id_medidor_activo,ip_fija,folder' +
             ' FROM A08_Medidor_Activo ' +
             'WHERE fecha_baja IS NULL AND id_medidor_activo = 10293') # Quitar 'and id_medidor_activo...'  MJF 20230615 
    # Ejecucion de consulta
    idMed_ipFija_folder = getDBdata(parametro, tabla, logFile=logFile,
                                    Query=Query)
    logSave(logFile,
            'Datos obtenido de la base de datos de manera exitosa',
            debug)

    idMed = []
    Ips = []  # Lista de Ips
    folders = []  # Lista de modelos
    for x in idMed_ipFija_folder:
        idMed.append(x[0])
        Ips.append(x[1])
        folders.append(x[2])

    idTipoMed = []
    for x in idMed:
        parametro = 'id_modelo'
        tabla = 'A05_Medidor'
        param_cond = 'id_medidor'
        valor_cond = "'" + str(x) + "'"
        idType = getDBdata(parametro, tabla,
                           logFile=logFile)

        idTipoMed.append(idType[0][0])

    modelos = []
    for x in idTipoMed:
        parametro = 'nombre_corto'
        tabla = 'A03_Modelo_Medidor'
        param_cond = 'id_modelo_medidor'
        valor_cond = "'" + str(x) + "'"

        mods = getDBdata(parametro, tabla, logFile=logFile)
        modelos.append(mods[0][0])

    idsClientes = []
    idsPlantas = []
    # Con el id de medidor, obtener id de cliente e id de planta
    for x in idMed:
        parametro = 'id_cliente,id_planta'
        tabla = 'A08_Medidor_Activo'
        param_cond = 'id_medidor_activo'
        valor_cond = "'" + str(x) + "'"
        # Lista de ids de cliente y de planta
        idCliente_idPlanta = getDBdata(parametro, tabla, param_cond,
                                       valor_cond, logFile=logFile)
        idsClientes.append(idCliente_idPlanta[0][0])
        idsPlantas.append(idCliente_idPlanta[0][1])

    # Con los id de clientes y de plantas, obtener los nombres de estos
    clientes = []
    for x in idsClientes:
        parametro = 'nombre_corto'
        tabla = 'A01_Cliente'
        param_cond = 'id_cliente'
        valor_cond = "'" + str(x) + "'"
        # Lista de ids de cliente y de planta
        listaClientes = getDBdata(parametro, tabla, param_cond,
                                  valor_cond, logFile=logFile)
        clientes.append(listaClientes[0][0])

    plantas = []
    for x in idsPlantas:
        parametro = 'nombre_corto'
        tabla = 'A02_Planta_Empresa'
        param_cond = 'id_planta'
        valor_cond = "'" + str(x) + "'"
        # Lista de ids de cliente y de planta
        listaPlantas = getDBdata(parametro, tabla, param_cond,
                                 valor_cond, logFile=logFile)
        plantas.append(listaPlantas[0][0])

    zipper = zip(Ips, modelos)
    typeDict = dict(zipper)  # Diccionario de Ips y sus modelos

    zipper = zip(Ips, folders)
    folderDict = dict(zipper)

    zipper = zip(Ips, clientes)
    clientesDict = dict(zipper)

    zipper = zip(Ips, plantas)
    plantasDict = dict(zipper)

    logSave(logFile, 'Ips y modelos definidos', debug)

except Exception:
    debug = '1'
    logSave(logFile, 'Fallo en conexion a base de datos.', debug)
    logSave(logFile, 'Inicia conexion con archivo XLSX', debug)
    # °°°°°Conexion a archivo de Ips y medidores definidos en archivo Excel°°°
    os.chdir("C:\\Users\\" + getpass.getuser() + "\\Desktop" +
             "\\Transferencia")
    wb = openpyxl.load_workbook('Medidores_IPs.xlsx')
    first_sheet = wb.sheetnames
    first_sheet = str(first_sheet[0])

    worksheet = wb[first_sheet]
    # Extraer lista de Ips
    Ips = []
    for row in range(2, worksheet.max_row + 1):
        for column in "A":
            cell_name = "{}{}".format(column, row)
            worksheet[cell_name].value
            Ips.append(worksheet[cell_name].value)

    logSave(logFile, 'IPs obtenidas', debug)

    folders = []
    for row in range(2, worksheet.max_row + 1):
        for column in "C":
            cell_name = "{}{}".format(column, row)
            worksheet[cell_name].value
            folders.append(worksheet[cell_name].value)

    logSave(logFile, 'Nombre de folders obtenidos', debug)

    # Extraer lista de modelos de medidor
    modelos = []
    for row in range(2, worksheet.max_row + 1):
        for column in "D":
            cell_name = "{}{}".format(column, row)
            worksheet[cell_name].value
            modelos.append(worksheet[cell_name].value)

    logSave(logFile, 'Modelos obtenidos', debug)

    clientes = []
    for row in range(2, worksheet.max_row + 1):
        for column in "E":
            cell_name = "{}{}".format(column, row)
            worksheet[cell_name].value
            clientes.append(worksheet[cell_name].value)

    logSave(logFile, 'Clientes obtenidos', debug)

    plantas = []
    for row in range(2, worksheet.max_row + 1):
        for column in "F":
            cell_name = "{}{}".format(column, row)
            worksheet[cell_name].value
            plantas.append(worksheet[cell_name].value)

    logSave(logFile, 'Plantas obtenidas', debug)

    zipper = zip(Ips, modelos)
    typeDict = dict(zipper)  # Diccionario de Ips y sus modelos

    zipper = zip(Ips, folders)
    folderDict = dict(zipper)   # Diccionario de Ips y sus folders

    zipper = zip(Ips, clientes)
    clientesDict = dict(zipper)  # Diccionario de Ips y sus clientes

    zipper = zip(Ips, plantas)
    plantasDict = dict(zipper)  # Diccionario de Ips y sus plantas

    logSave(logFile, 'Ips y diccionarios definidos', debug)

# °°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°
# Obtener los que son PureBB y G44X0 por separado
ipsBB = []
ipsG4 = []

for key in typeDict.keys():
    if typeDict[key] == 'PureBB':
        ipsBB.append(key)
    elif typeDict[key] == 'G44X0':
        ipsG4.append(key)

logSave(logFile, 'Clasificacion de Ips por modelo realizada', debug)
# °°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°
logSave(logFile, 'Inicia extraccion de PQZs para medidores PureBB', debug)

ExtraccionBB = PureBBPQZs
ExtraccionBB.RecoverProcess(ipsBB, logFile, debug, folderDict, clientesDict,
                            plantasDict)
# °°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°°
# logSave(logFile, 'Inicia extraccion de PQZs para medidores G44X0')

ExtraccionG4 = G44X0PQZs
ExtraccionG4.RecoverProcess(ipsG4, logFile, debug, folderDict, clientesDict,
                            plantasDict)
